# -*- coding: utf-8 -*-
import base64
import io
import re
import unicodedata
import zipfile
from datetime import date, datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None

GENDER_MAP = {"M": "m", "F": "f"}

# Mots-cles (normalises : minuscules, sans accents) recherches dans chaque
# cellule d'en-tete pour identifier la colonne, quelle que soit sa position
# exacte ou sa formulation precise dans le fichier source.
COLUMN_KEYWORDS = {
    "matricule": "matricule",
    "prenom": "prenom",
    "nom": "nom",
    "sexe": "sexe",
    "date nais": "date_naissance",
    "lieu naissance": "lieu_naissance",
    "salle de classe": "salle_classe",
    "est inscrit": "est_inscrit",
    "reduction": "reduction",
    "date inscription": "date_inscription",
    "montant": "montant_paye",
    "ecole": "ecole_origine",
}


def _normalize(text):
    if text is None:
        return ""
    text = unicodedata.normalize("NFKD", str(text))
    return "".join(c for c in text if not unicodedata.combining(c)).strip().lower()


class OpStudentImportWizard(models.TransientModel):
    _name = "op.student.import.wizard"
    _description = "Import en masse des eleves (fichier Excel)"

    state = fields.Selection(
        [("upload", "Fichier"), ("preview", "Apercu"), ("done", "Termine")],
        default="upload", required=True)

    file_data = fields.Binary(string="Fichier Excel (.xlsx)")
    file_name = fields.Char(string="Nom du fichier")
    academic_year_id = fields.Many2one(
        "op.academic.year", string="Annee academique",
        default=lambda self: self.env.user.current_academic_year_id, required=True)
    migration_fee_type_id = fields.Many2one(
        "op.fee.type", string="Type de frais pour le solde migre",
        help="Utilise pour comptabiliser, pour chaque eleve, la part du "
             "'montant deja paye' qui depasse les frais deja configures sur "
             "sa classe (excedent, ou classe sans frais configures).")
    payment_journal_id = fields.Many2one(
        "account.journal", string="Journal de paiement",
        domain=[("type", "in", ("bank", "cash"))],
        default=lambda self: self.env["account.journal"].search(
            [("type", "in", ("bank", "cash")),
             ("company_id", "=", self.env.company.id)], limit=1))

    preview_line_ids = fields.One2many(
        "op.student.import.wizard.line", "wizard_id", string="Apercu des eleves")
    has_payments = fields.Boolean(compute="_compute_has_payments")

    result_summary = fields.Text(string="Resultat", readonly=True)

    @api.depends("preview_line_ids.montant_paye")
    def _compute_has_payments(self):
        for wizard in self:
            wizard.has_payments = any(wizard.preview_line_ids.mapped("montant_paye"))

    # ------------------------------------------------------------------
    # Lecture du fichier
    # ------------------------------------------------------------------
    @staticmethod
    def _sanitize_xlsx(raw):
        """Certains fichiers .xlsx (resaves par des outils tiers / vieux
        modeles) contiennent un <cellStyle> sans attribut 'name' dans
        xl/styles.xml : invalide selon le schema OOXML mais tolere par
        Excel. openpyxl applique le schema strictement et leve un
        TypeError au chargement - on corrige l'attribut manquant avant de
        passer le fichier a openpyxl plutot que de toucher la librairie
        installee sur le serveur."""
        try:
            with zipfile.ZipFile(io.BytesIO(raw)) as zin:
                if "xl/styles.xml" not in zin.namelist():
                    return raw
                styles_xml = zin.read("xl/styles.xml")
        except zipfile.BadZipFile:
            return raw

        patched = re.sub(
            rb"<cellStyle(?![^>]*\sname=)([^>]*)/>",
            rb'<cellStyle name="Normal"\1/>',
            styles_xml,
        )
        if patched == styles_xml:
            return raw

        out = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(raw)) as zin, \
                zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = patched if item.filename == "xl/styles.xml" else zin.read(item.filename)
                zout.writestr(item, data)
        return out.getvalue()

    def _detect_header(self, sheet):
        for row in sheet.iter_rows(values_only=True):
            normalized = [_normalize(v) for v in row]
            if any("matricule" in v for v in normalized):
                mapping = {}
                for col_index, header in enumerate(normalized):
                    for keyword, key in COLUMN_KEYWORDS.items():
                        if keyword in header and key not in mapping.values():
                            mapping[col_index] = key
                            break
                return mapping
        return None

    def _read_rows(self):
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_(
                "La librairie Python 'openpyxl' n'est pas installee sur le "
                "serveur (pip install openpyxl)."))
        if not self.file_data:
            raise UserError(_("Veuillez selectionner un fichier Excel (.xlsx)."))

        raw = self._sanitize_xlsx(base64.b64decode(self.file_data))
        book = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        sheet = book.active
        column_map = self._detect_header(sheet)
        if not column_map or "matricule" not in column_map.values():
            raise UserError(_(
                "Impossible de trouver la ligne d'en-tete (une colonne "
                "'Matricule' est requise) dans le fichier."))

        rows = []
        header_found = False
        for row in sheet.iter_rows(values_only=True):
            if not header_found:
                normalized = [_normalize(v) for v in row]
                if any("matricule" in v for v in normalized):
                    header_found = True
                continue
            if not any(v not in (None, "") for v in row):
                continue
            data = {key: row[idx] if idx < len(row) else None
                    for idx, key in column_map.items()}
            if not data.get("matricule"):
                continue
            rows.append(data)
        return rows

    # ------------------------------------------------------------------
    # Conversions
    # ------------------------------------------------------------------
    @staticmethod
    def _to_float(value):
        if value in (None, ""):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(" ", "").replace(",", "."))
        except ValueError:
            return 0.0

    @staticmethod
    def _to_date(value):
        if not value:
            return False
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()
        except ValueError:
            return False

    def _get_cameroun_id(self):
        country = self.env.ref("base.cm", raise_if_not_found=False)
        return country.id if country else False

    # ------------------------------------------------------------------
    # Classe : recherche exacte, sinon creation automatique si le niveau
    # peut etre devine depuis le nom (ex: "CE2 C" -> niveau ce2).
    # ------------------------------------------------------------------
    def _guess_level(self, name, levels_sorted):
        normalized = _normalize(name)
        for level in levels_sorted:
            if normalized.startswith(_normalize(level.name)):
                return level.code
        return False

    def _get_or_create_classe(self, name, levels_sorted, classe_cache):
        """levels_sorted et classe_cache sont prepares une seule fois par
        import (cf. action_preview) plutot que re-recherches a chaque ligne :
        un fichier de plusieurs centaines d'eleves reutilise le plus souvent
        une poignee de classes/niveaux, et repeter ces recherches identiques
        des centaines de fois ajoute un cout inutile (jusqu'a declencher le
        depassement de limite CPU du worker sur de gros fichiers)."""
        if name in classe_cache:
            return classe_cache[name], False
        Classe = self.env["op.classe"]
        classe = Classe.search([
            ("name", "=", name),
            ("academic_year_id", "=", self.academic_year_id.id),
        ], limit=1)
        if classe:
            classe_cache[name] = classe
            return classe, False
        level = self._guess_level(name, levels_sorted)
        if not level:
            return Classe, False
        classe = Classe.create({
            "name": name,
            "level": level,
            "academic_year_id": self.academic_year_id.id,
        })
        self._copy_fee_amounts_from_sibling_classe(classe, level)
        classe_cache[name] = classe
        return classe, True

    def _copy_fee_amounts_from_sibling_classe(self, classe, level):
        """Une classe nouvellement creee demarre avec les frais standards de
        l'ecole a 0 (cf. op.classe._ensure_default_fee_lines), donc aucun
        frais ne serait genere pour les eleves importes dans cette classe.
        Reprend ici les montants deja configures sur une autre classe du
        meme niveau pour la meme annee (s'il en existe une), pour que les
        frais se generent automatiquement sans configuration manuelle prealable."""
        reference = self.env["op.classe"].search([
            ("level", "=", level),
            ("academic_year_id", "=", self.academic_year_id.id),
            ("id", "!=", classe.id),
            ("fee_line_ids.amount", ">", 0),
        ], limit=1)
        if not reference:
            return
        existing_by_type = {l.fee_type_id.id: l for l in classe.fee_line_ids}
        new_vals = []
        for ref_line in reference.fee_line_ids:
            if ref_line.amount <= 0:
                continue
            line = existing_by_type.get(ref_line.fee_type_id.id)
            if line:
                line.amount = ref_line.amount
            else:
                new_vals.append({
                    "classe_id": classe.id,
                    "fee_type_id": ref_line.fee_type_id.id,
                    "amount": ref_line.amount,
                })
        if new_vals:
            self.env["op.classe.fee"].create(new_vals)

    def _line_to_student_vals(self, line):
        vals = {
            "first_name": line.first_name or "",
            "last_name": line.last_name or "",
            "name": " ".join(filter(None, [line.last_name, line.first_name])) or "/",
            "gender": line.gender or "m",
            "birth_date": line.birth_date,
            "birth_place": line.birth_place or "",
            "ecole_origine": line.ecole_origine or "",
            "discount_amount": line.reduction,
            "nationality": self._get_cameroun_id(),
        }
        return {k: v for k, v in vals.items() if v not in (False, "")}

    def _reopen(self):
        return {
            "type": "ir.actions.act_window",
            "res_model": "op.student.import.wizard",
            "view_mode": "form",
            "res_id": self.id,
            "target": "current",
        }

    # ------------------------------------------------------------------
    # Etape 1 : apercu / edition
    # ------------------------------------------------------------------
    def action_preview(self):
        self.ensure_one()
        rows = self._read_rows()
        if not rows:
            raise UserError(_("Aucune ligne exploitable trouvee dans le fichier."))

        levels_sorted = self.env["op.education.level"].search([])
        levels_sorted = sorted(levels_sorted, key=lambda lvl: -len(lvl.name))
        classe_cache = {}

        vals_list = []
        for index, row in enumerate(rows, start=2):
            matricule = str(row.get("matricule") or "").strip()
            if not matricule:
                continue

            salle = str(row.get("salle_classe") or "").strip()
            classe_id = False
            classe_auto_created = False
            warning = False
            if salle:
                classe, was_created = self._get_or_create_classe(salle, levels_sorted, classe_cache)
                if classe:
                    classe_id = classe.id
                    classe_auto_created = was_created
                else:
                    warning = _("Classe '%s' introuvable - a choisir manuellement.") % salle

            vals_list.append({
                "wizard_id": self.id,
                "row_index": index,
                "matricule": matricule,
                "first_name": str(row.get("prenom") or "").strip().title(),
                "last_name": str(row.get("nom") or "").strip().title(),
                "gender": GENDER_MAP.get(_normalize(row.get("sexe"))[:1].upper(), "m"),
                "birth_date": self._to_date(row.get("date_naissance")),
                "birth_place": str(row.get("lieu_naissance") or "").strip(),
                "ecole_origine": str(row.get("ecole_origine") or "").strip(),
                "est_inscrit": _normalize(row.get("est_inscrit")).startswith("o"),
                "classe_id": classe_id,
                "classe_texte": salle,
                "classe_auto_created": classe_auto_created,
                "reduction": self._to_float(row.get("reduction")),
                "date_inscription": self._to_date(row.get("date_inscription")),
                "montant_paye": self._to_float(row.get("montant_paye")),
                "warning": warning,
            })

        self.preview_line_ids.unlink()
        self.env["op.student.import.wizard.line"].create(vals_list)
        self.state = "preview"
        return self._reopen()

    def action_back_to_upload(self):
        self.ensure_one()
        self.preview_line_ids.unlink()
        self.state = "upload"
        self.result_summary = False
        return self._reopen()

    # ------------------------------------------------------------------
    # Etape 2 : import effectif a partir de l'apercu (eventuellement edite)
    # ------------------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        if self.state != "preview":
            raise UserError(_("Veuillez d'abord generer l'apercu du fichier."))
        if not self.preview_line_ids:
            raise UserError(_("Aucune ligne a importer."))

        Student = self.env["op.student"]
        created = updated = fees_settled = 0
        errors = []

        for line in self.preview_line_ids:
            matricule = (line.matricule or "").strip()
            if not matricule:
                continue

            vals = self._line_to_student_vals(line)
            # Le matricule n'est plus unique que par annee academique (cf.
            # op.student._unique_gr_no) : un meme matricule peut legitimement
            # reapparaitre sur le dossier d'une AUTRE annee (cloture
            # d'annee). On precise donc explicitement l'annee de CE wizard -
            # academic_year_all=True pour ignorer l'annee courante de
            # l'utilisateur, potentiellement differente de celle importee ici.
            student = Student.with_context(academic_year_all=True).search([
                ("gr_no", "=", matricule),
                ("academic_year_id", "=", self.academic_year_id.id),
            ], limit=1)
            if student:
                student.write(vals)
                updated += 1
            else:
                vals["gr_no"] = matricule
                student = Student.create(vals)
                created += 1

            if line.classe_id and student.classe_id != line.classe_id:
                student.classe_id = line.classe_id.id

            if line.montant_paye > 0:
                if not self.migration_fee_type_id or not self.payment_journal_id:
                    errors.append(_(
                        "Ligne %(line)s (%(matricule)s) : montant deja paye renseigne "
                        "mais 'Type de frais pour le solde migre' / 'Journal de "
                        "paiement' non renseignes - paiement ignore."
                    ) % {"line": line.row_index, "matricule": matricule})
                else:
                    try:
                        self._apply_migration_payment(student, line.montant_paye, line)
                        fees_settled += 1
                    except Exception as exc:  # noqa: BLE001 - on isole l'erreur par ligne
                        errors.append(_(
                            "Ligne %(line)s (%(matricule)s) : echec de l'enregistrement "
                            "du paiement (%(error)s)."
                        ) % {"line": line.row_index, "matricule": matricule, "error": str(exc)})

        classes_created = sum(self.preview_line_ids.mapped("classe_auto_created"))
        self.result_summary = _(
            "%(created)s eleve(s) cree(s), %(updated)s mis a jour, "
            "%(classes)s classe(s) creee(s) automatiquement, "
            "%(fees)s solde(s) de migration enregistre(s).\n\n%(errors)s"
        ) % {
            "created": created, "updated": updated, "classes": classes_created,
            "fees": fees_settled,
            "errors": "\n".join(errors) if errors else _("Aucune erreur."),
        }
        self.state = "done"
        return self._reopen()

    def _apply_migration_payment(self, student, amount, line):
        """Affecte le montant deja paye en priorite au frais d'inscription
        (necessaire pour que l'eleve soit considere "Inscrit"), puis aux
        autres frais deja generes sur sa classe, puis a un frais de
        migration dedie pour le reliquat eventuel."""
        remaining = amount
        outstanding = student.fee_ids.filtered(
            lambda f: f.state == "posted" and f.payment_state != "paid"
            and f.academic_year_id == self.academic_year_id)
        registration = outstanding.filtered(lambda f: f.fee_type_id.is_registration_fee)
        others = (outstanding - registration).sorted("date")

        for fee in registration + others:
            if remaining <= 0:
                break
            pay_amount = min(remaining, fee.amount_residual)
            if pay_amount <= 0:
                continue
            self._register_payment(fee, pay_amount, line)
            remaining -= pay_amount

        if remaining > 0:
            sale_journal = self.env["account.journal"].search(
                [("type", "=", "sale"), ("company_id", "=", self.env.company.id)], limit=1)
            fee = self.env["op.student.fee"].create({
                "student_id": student.id,
                "fee_type_id": self.migration_fee_type_id.id,
                "academic_year_id": self.academic_year_id.id,
                "date": line.date_inscription or fields.Date.context_today(self),
                "amount": remaining,
                "journal_id": sale_journal.id,
            })
            fee.action_post()
            self._register_payment(fee, remaining, line)

    def _register_payment(self, fee, amount, line):
        fee.ensure_one()
        register = self.env["account.payment.register"].with_context(
            active_model="account.move", active_ids=[fee.move_id.id],
        ).create({
            "journal_id": self.payment_journal_id.id,
            "amount": amount,
            "payment_date": line.date_inscription or fields.Date.context_today(self),
        })
        register._create_payments()
